"""Tests for scripts/greenwood_migrate.py."""
import json
from pathlib import Path

import openpyxl
import pytest

from scripts.greenwood_migrate import (
    sanitize_dir_name,
    clean_tier1_name,
    load_db_mapping,
    scan_local,
    plan_migration,
    execute_moves,
)


def create_realistic_db(path):
    """Create a Master DB + Tier1 sheet layout matching the real HealthcareIntel DB."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Master DB")
    ws.append(["HealthcareIntel Database — Master List"])
    ws.append(["Stats line"])
    ws.append([])
    ws.append(["#", "Company", "Ticker", "Exchange", "Mkt Cap (USD)",
               "Tier 1", "Tier 2", "Focus / Notes", "NEW"])
    ws.append(["Sector 1: Biopharma"])
    ws.append([1, "AbbVie", "ABBV", "NYSE", "$300B",
               "1. Biopharma", "1.1 Oncology", "Oncology leader", ""])
    ws.append([2, "10x Genomics", "TXG", "NASDAQ", "$1B",
               "4. Biologics Tools & Services", "4.2 Spatial Biology",
               "Spatial genomics", ""])
    ws.append([3, "Hologic", "HOLX", "NASDAQ", "$20B",
               "7. IVD", "7.1 Women's Health Dx", "Women's health", ""])
    ws.append([4, "Align Technology", "ALGN", "NASDAQ", "$20B",
               "9. Dentistry", "9.1 Clear Aligners", "Invisalign", ""])

    # Tier1 sheet with same 3-row preamble
    ws2 = wb.create_sheet("1. Biopharma")
    ws2.append(["Sector 1: Biopharma — 1 company"])
    ws2.append([])
    ws2.append(["#", "Company", "Ticker", "Exchange", "Mkt Cap (USD)",
                "Sub-sector", "Focus / Notes", "NEW"])
    ws2.append(["1.1 Oncology"])
    ws2.append([1, "AbbVie", "ABBV", "NYSE", "$300B",
                "1.1 Oncology", "Oncology leader", ""])

    wb.save(path)
    return path


@pytest.fixture
def realistic_db(tmp_path):
    return create_realistic_db(tmp_path / "HealthcareIntel_Database_20260412.xlsx")


@pytest.fixture
def local_tree(tmp_path):
    """Simulate user's Greenwood local tree with old sector names."""
    root = tmp_path / "Earnings"
    fy = root / "2025_FY"

    # ABBV already in Biopharma (no-op expected)
    (fy / "Biopharma" / "ABBV").mkdir(parents=True)
    (fy / "Biopharma" / "ABBV" / "ABBV_2025FY_Transcript.txt").write_text("x" * 2000)

    # TXG in Diagnostics_LifeSci (should move to Biologics_Tools_and_Services)
    (fy / "Diagnostics_LifeSci" / "TXG").mkdir(parents=True)
    (fy / "Diagnostics_LifeSci" / "TXG" / "TXG_2025FY_Transcript.txt").write_text("x" * 2000)

    # HOLX in Diagnostics_LifeSci (should move to IVD)
    (fy / "Diagnostics_LifeSci" / "HOLX").mkdir(parents=True)
    (fy / "Diagnostics_LifeSci" / "HOLX" / "HOLX_2025FY_Transcript.txt").write_text("x" * 2000)

    # ALGN in Dental (should move to Dentistry)
    (fy / "Dental" / "ALGN").mkdir(parents=True)
    (fy / "Dental" / "ALGN" / "ALGN_2025FY_Transcript.txt").write_text("x" * 2000)

    return root


class TestCleanTier1Name:
    def test_no_prefix(self):
        assert clean_tier1_name("Biopharma") == "Biopharma"

    def test_strip_number_dot(self):
        assert clean_tier1_name("1. Biopharma") == "Biopharma"

    def test_strip_number_space(self):
        assert clean_tier1_name("1 Biopharma") == "Biopharma"

    def test_strip_sector_prefix(self):
        assert clean_tier1_name("Sector 1: Biopharma") == "Biopharma"

    def test_empty(self):
        assert clean_tier1_name("") == ""


class TestSanitizeDir:
    def test_basic(self):
        assert sanitize_dir_name("Biopharma") == "Biopharma"

    def test_strips_number_prefix(self):
        assert sanitize_dir_name("1. Biopharma") == "Biopharma"
        assert sanitize_dir_name("9 Dentistry") == "Dentistry"

    def test_ampersand(self):
        assert sanitize_dir_name("Biologics Tools & Services") == "Biologics_Tools_and_Services"

    def test_ampersand_with_prefix(self):
        assert sanitize_dir_name("4. Biologics Tools & Services") == "Biologics_Tools_and_Services"

    def test_spaces(self):
        assert sanitize_dir_name("Healthcare Services") == "Healthcare_Services"


class TestFirstOccurrenceWins:
    def test_ticker_in_multiple_sectors(self, tmp_path):
        """When a ticker appears in multiple Tier 1 rows, keep the first."""
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Master DB")
        ws.append(["HealthcareIntel Database"])
        ws.append(["Stats"])
        ws.append([])
        ws.append(["#", "Company", "Ticker", "Exchange", "Mkt Cap (USD)",
                   "Tier 1", "Tier 2", "Focus / Notes", "NEW"])
        ws.append(["Sector 1: Biopharma"])
        ws.append([1, "AbbVie", "ABBV", "NYSE", "$300B",
                   "1. Biopharma", "1.1 Oncology", "Pharma", ""])
        ws.append(["Sector 2: MedTech"])
        # Duplicate ABBV under MedTech — should NOT override Biopharma
        ws.append([2, "AbbVie (devices)", "ABBV", "NYSE", "$300B",
                   "2. MedTech", "2.5 Aesthetics", "Allergan aesthetics", ""])
        path = tmp_path / "db.xlsx"
        wb.save(path)

        mapping = load_db_mapping(str(path))
        assert mapping["ABBV"]["tier1"] == "Biopharma"  # NOT MedTech
        assert mapping["ABBV"]["sub_sector"] == "1.1 Oncology"


class TestLoadDbMapping:
    def test_uses_master_db(self, realistic_db):
        mapping = load_db_mapping(str(realistic_db))
        assert "ABBV" in mapping
        assert mapping["ABBV"]["tier1"] == "Biopharma"
        assert mapping["TXG"]["tier1"] == "Biologics Tools & Services"
        assert mapping["HOLX"]["tier1"] == "IVD"
        assert mapping["ALGN"]["tier1"] == "Dentistry"

    def test_sub_sector_from_tier2(self, realistic_db):
        mapping = load_db_mapping(str(realistic_db))
        assert mapping["TXG"]["sub_sector"] == "4.2 Spatial Biology"
        assert mapping["ALGN"]["sub_sector"] == "9.1 Clear Aligners"


class TestScanLocal:
    def test_finds_all_tickers(self, local_tree):
        entries = scan_local(local_tree / "2025_FY")
        tickers = [e["ticker"] for e in entries]
        assert "ABBV" in tickers
        assert "TXG" in tickers
        assert "HOLX" in tickers
        assert "ALGN" in tickers

    def test_empty_dir(self, tmp_path):
        entries = scan_local(tmp_path)
        assert entries == []


class TestPlanMigration:
    def test_cross_sector_moves(self, realistic_db, local_tree):
        mapping = load_db_mapping(str(realistic_db))
        local = scan_local(local_tree / "2025_FY")
        moves, unmapped = plan_migration(local, mapping)

        abbv_move = [m for m in moves if m["ticker"] == "ABBV"][0]
        assert abbv_move["op"] == "no_op"

        txg_move = [m for m in moves if m["ticker"] == "TXG"][0]
        assert txg_move["op"] == "move"
        assert "Biologics_Tools_and_Services" in txg_move["to"]

        holx_move = [m for m in moves if m["ticker"] == "HOLX"][0]
        assert holx_move["op"] == "move"
        assert "IVD" in holx_move["to"]

        algn_move = [m for m in moves if m["ticker"] == "ALGN"][0]
        assert algn_move["op"] == "move"
        assert "Dentistry" in algn_move["to"]

        assert len(unmapped) == 0

    def test_unmapped_ticker(self, realistic_db, tmp_path):
        root = tmp_path / "Earnings"
        fy = root / "2025_FY"
        (fy / "Biopharma" / "UNKNOWN").mkdir(parents=True)
        (fy / "Biopharma" / "UNKNOWN" / "UNKNOWN_2025FY_Transcript.txt").write_text("x")

        mapping = load_db_mapping(str(realistic_db))
        local = scan_local(fy)
        moves, unmapped = plan_migration(local, mapping)

        assert len(unmapped) == 1
        assert unmapped[0]["ticker"] == "UNKNOWN"


class TestExecuteMoves:
    def test_dry_run(self, realistic_db, local_tree):
        mapping = load_db_mapping(str(realistic_db))
        local = scan_local(local_tree / "2025_FY")
        moves, unmapped = plan_migration(local, mapping)

        stats = execute_moves(moves, unmapped, local_tree / "2025_FY", dry_run=True)

        # Files should NOT have moved
        assert (local_tree / "2025_FY" / "Diagnostics_LifeSci" / "TXG").exists()
        assert not (local_tree / "2025_FY" / "Biologics_Tools_and_Services" / "TXG").exists()

    def test_execute_moves(self, realistic_db, local_tree):
        mapping = load_db_mapping(str(realistic_db))
        local = scan_local(local_tree / "2025_FY")
        moves, unmapped = plan_migration(local, mapping)

        stats = execute_moves(moves, unmapped, local_tree / "2025_FY", dry_run=False)

        assert stats["moved"] >= 3  # TXG, HOLX, ALGN
        assert stats["no_op"] == 1  # ABBV

        # Files should be in new locations
        assert (local_tree / "2025_FY" / "Biologics_Tools_and_Services" / "TXG").exists()
        assert (local_tree / "2025_FY" / "IVD" / "HOLX").exists()
        assert (local_tree / "2025_FY" / "Dentistry" / "ALGN").exists()

        # ABBV stays in Biopharma
        assert (local_tree / "2025_FY" / "Biopharma" / "ABBV").exists()

        # Old empty sector dirs should be cleaned up
        assert not (local_tree / "2025_FY" / "Diagnostics_LifeSci").exists()
        assert not (local_tree / "2025_FY" / "Dental").exists()
